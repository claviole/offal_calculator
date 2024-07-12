import re
from collections import defaultdict
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import sys
import os

def parse_report(file_path):
    with open(file_path, 'r') as file:
        lines = file.readlines()

    monthly_data = defaultdict(lambda: {'coils': [], 'shifts': 0, 'scrap_weight': 0})
    total_shifts = 0
    total_scrap_weight = 0.0
    total_scrap_lbs = 0.0  # To accumulate ScrapLbs for Offal calculation

    current_month = None
    line_number = None
    dates = []
    coil_weights = defaultdict(float)  # Initialize as defaultdict to accumulate weights
    processed_coils = set()  # Track processed coils
    skip_until_new_date = False  # Flag to skip lines until a new date is found
    found_scrap_lbs = False  # Flag to stop looking for scrap lbs after it is found

    i = 0
    while i < len(lines):
        line = lines[i]
        
        # Match lines with the date format m/d/yy or m/d/yyyy, allowing for leading whitespace
        date_match = re.match(r'\s*(\d{1,2}/\d{1,2}/\d{2,4})', line)
        if date_match:
            skip_until_new_date = False  # Reset the flag when a new date is found
            found_scrap_lbs = False  # Reset the flag for scrap lbs
            date_str = date_match.group(1)
            try:
                date = datetime.strptime(date_str, '%m/%d/%y').date()
            except ValueError:
                date = datetime.strptime(date_str, '%m/%d/%Y').date()
            current_month = date.strftime('%Y-%m')
            dates.append(date)
            
            # Extract shift number (next value after the date)
            parts = line.split()
            if len(parts) >= 3:
                shift_number = int(parts[1])
                line_number = parts[2]
                total_shifts += 1
                monthly_data[current_month]['shifts'] += 1
            
            # Extract coil weight (second to last value in the line)
            coil_weight_match = re.findall(r'\d+\.\d+|\d+', line)
            if coil_weight_match:
                coil_weight = float(coil_weight_match[-2])
                coil_weights[current_month] += coil_weight  # Accumulate coil weight
            
            # Extract coil number (second 6-digit number in the line)
            coil_number_match = re.findall(r'\b\d{6}\b', line)
            if len(coil_number_match) >= 2:
                coil_number = coil_number_match[1]
            
            # Check if the coil has already been processed
            if coil_number in processed_coils:
                skip_until_new_date = True  # Set the flag to skip lines until a new date is found
                i += 1
                continue
            
            # Mark the coil as processed
            processed_coils.add(coil_number)
            
            # Initialize a new coil entry
            monthly_data[current_month]['coils'].append({
                'coil_number': coil_number,
                'scrap_weight': 0.0,
                'scrap_lbs': 0.0,
                'coil_weight': coil_weight,  # Store the coil weight
                'excess_scrap': 0.0,  # Initialize excess_scrap
                'unaccounted_scrap': 0.0,  # Initialize unaccounted_scrap
                'offal': 0.0  # Initialize offal
            })
        elif "Total" in line:
            # If "Total" is encountered, stop looking for scrap weight and scrap lbs
            i += 1
            continue
        elif skip_until_new_date:
            # Skip lines until a new date is found
            i += 1
            continue
        else:
            # Accumulate scrap weight
            scrap_match = re.search(r'(\d+) Lbs', line)
            if scrap_match and monthly_data[current_month]['coils']:
                scrap_weight = float(scrap_match.group(1))
                total_scrap_weight += scrap_weight
                monthly_data[current_month]['coils'][-1]['scrap_weight'] += scrap_weight
                monthly_data[current_month]['scrap_weight'] += scrap_weight
            
                # Calculate excess scrap and ensure it is not negative
                coil_weight = monthly_data[current_month]['coils'][-1]['coil_weight']
                excess_scrap = scrap_weight - (coil_weight * 0.02)
                if excess_scrap < 0:
                    excess_scrap = 0.0
                monthly_data[current_month]['coils'][-1]['excess_scrap'] = excess_scrap
            
            # Look for scrap lbs only if it hasn't been found yet
            if not found_scrap_lbs:
                scrap_lbs_match = re.search(r'% Scrap\s+(\d+)# Scrap', line)
                if scrap_lbs_match and monthly_data[current_month]['coils']:
                    scrap_lbs = float(scrap_lbs_match.group(1))
                    total_scrap_lbs += scrap_lbs
                    monthly_data[current_month]['coils'][-1]['scrap_lbs'] += scrap_lbs
                    
                    # Calculate unaccounted scrap and ensure it is not negative
                    unaccounted_scrap = scrap_lbs - (coil_weight * 0.10)
                    if unaccounted_scrap < 0:
                        unaccounted_scrap = 0.0
                    monthly_data[current_month]['coils'][-1]['unaccounted_scrap'] = unaccounted_scrap

                    monthly_data[current_month]['coils'][-1]['scrap_lbs'] -= unaccounted_scrap
                    found_scrap_lbs = True  # Set the flag to stop looking for scrap lbs

                    # Calculate offal and ensure it is not negative
                    offal = scrap_lbs - scrap_weight
                    if offal < 0:
                        offal = 0.0
                    monthly_data[current_month]['coils'][-1]['offal'] = offal
                
        i += 1

    date_range = f"{min(dates).strftime('%Y-%m-%d')} to {max(dates).strftime('%Y-%m-%d')}" if dates else "No Date Range"
    return monthly_data, total_shifts, line_number, date_range, total_scrap_weight, coil_weights, total_scrap_lbs

def generate_report(monthly_data, total_shifts, total_scrap_weight, coil_weights, total_scrap_lbs):
    report = []
    total_hours = total_shifts * 7.25

    report.append(f"Total Shifts: {total_shifts}")

    report.append(f"Total Scrap Weight: {total_scrap_weight:.2f} Lbs")
    total_coil_weight = sum(coil_weights.values())
    excess_total_scrap = total_scrap_weight - (total_coil_weight * 0.02)
    if excess_total_scrap < 0:
        excess_total_scrap = 0.0
    report.append(f"Excess Total Scrap: {excess_total_scrap:.2f} Lbs")
    report.append(f"Total ScrapLbs: {total_scrap_lbs:.2f} Lbs")
    report.append(f"Offal: {max(total_scrap_lbs - total_scrap_weight, 0):.2f} Lbs")

    report.append("\nMonthly Data:")
    for month, info in sorted(monthly_data.items()):
        excess_scrap = max(info['scrap_weight'] - (coil_weights[month] * 0.02), 0.0)
        report.append(f"Month: {month}, Shifts: {info['shifts']}, Scrap Weight: {info['scrap_weight']:.2f} Lbs, Excess Scrap: {excess_scrap:.2f} Lbs")

    return "\n".join(report)

def generate_excel_report(monthly_data, total_shifts, line_number, date_range, total_scrap_weight, coil_weights, total_scrap_lbs, file_path):
    total_hours = total_shifts * 7.25
    total_coil_weight = sum(coil_weights.values())

    total_shifts_sum = 0
    total_scrap_weight_sum = 0.0
    total_excess_scrap_sum = 0.0
    total_scrap_lbs_sum = 0.0
    total_offal_sum = 0.0
    total_unaccounted_scrap_sum = 0.0

    with pd.ExcelWriter(file_path) as writer:
        for month, info in sorted(monthly_data.items()):
            month_data = {
                'Coil Number': [coil['coil_number'] for coil in info['coils']],
                'Scrap Weight': [coil['scrap_weight'] for coil in info['coils']],
                'Coil Weight': [coil['coil_weight'] for coil in info['coils']],
                'Excess Scrap': [max(coil['scrap_weight'] - (coil['coil_weight'] * 0.02), 0.0) for coil in info['coils']],
                'ScrapLbs': [coil['scrap_lbs'] for coil in info['coils']],
                'Unaccounted Scrap': [coil['unaccounted_scrap'] for coil in info['coils']],
                'Offal': [max(coil['scrap_lbs'] - coil['scrap_weight'], 0.0) for coil in info['coils']]
            }
            month_df = pd.DataFrame(month_data)
            month_df.to_excel(writer, sheet_name=month, index=False)
            # Add summary row
            summary_data = {
                'Coil Number': ['Total'],
                'Scrap Weight': [info['scrap_weight']],
                'Coil Weight': [sum(coil['coil_weight'] for coil in info['coils'])],
                'Excess Scrap': [sum(max(coil['scrap_weight'] - (coil['coil_weight'] * 0.02), 0.0) for coil in info['coils'])],
                'ScrapLbs': [sum(coil['scrap_lbs'] for coil in info['coils'])],
                'Unaccounted Scrap': [sum(coil['unaccounted_scrap'] for coil in info['coils'])],
                'Offal': [sum(max(coil['scrap_lbs'] - coil['scrap_weight'], 0.0) for coil in info['coils'])]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name=month, startrow=len(month_df) + 2, index=False)

            # Aggregate totals for the "Total" tab
            total_shifts_sum += info['shifts']
            total_scrap_weight_sum += info['scrap_weight']
            total_excess_scrap_sum += summary_data['Excess Scrap'][0]
            total_scrap_lbs_sum += summary_data['ScrapLbs'][0]
            total_offal_sum += summary_data['Offal'][0]
            total_unaccounted_scrap_sum += summary_data['Unaccounted Scrap'][0]

        overall_data = {
            'Total Shifts': [total_shifts_sum],
            'Total Scrap Weight': [total_scrap_weight_sum],
            'Excess Total Scrap': [total_excess_scrap_sum],
            'Total ScrapLbs': [total_scrap_lbs_sum],
            'Unaccounted Scrap': [total_unaccounted_scrap_sum],
            'Offal': [total_offal_sum]
        }
        overall_df = pd.DataFrame(overall_data)
        overall_df.to_excel(writer, sheet_name='Total', index=False)

    # Load the workbook and add Line # and Date Range to the first sheet
    workbook = load_workbook(file_path)
    worksheet_total = workbook['Total']
    worksheet_total['A1'] = f"Line: {line_number}"
    worksheet_total['A2'] = f"Date Range: {date_range}"

    # Apply formatting to Total Data sheet
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    for col in worksheet_total.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet_total.column_dimensions[column].width = adjusted_width

    for cell in worksheet_total["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Apply formatting to Monthly Data sheets
    for month in sorted(monthly_data.keys()):
        worksheet_month = workbook[month]
        for col in worksheet_month.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet_month.column_dimensions[column].width = adjusted_width

        for cell in worksheet_month["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

    workbook.save(file_path)

if __name__ == "__main__":
    import os
    import sys

    # Check if a file path is provided as a command-line argument
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        # Prompt the user for the file path
        file_path = input("Please enter the path to the text file: ").strip()
    
    if not os.path.isfile(file_path):
        print(f"File not found: {file_path}")
    else:
        monthly_data, total_shifts, line_number, date_range, total_scrap_weight, coil_weights, total_scrap_lbs = parse_report(file_path)
        report = generate_report(monthly_data, total_shifts, total_scrap_weight, coil_weights, total_scrap_lbs)
        print(report)

        excel_file_path = f'Line_{line_number}_{date_range.replace(" ", "_").replace(":", "-")}.xlsx'  # Replace with your desired Excel file path
        generate_excel_report(monthly_data, total_shifts, line_number, date_range, total_scrap_weight, coil_weights, total_scrap_lbs, excel_file_path)
        print(f"Excel report generated: {excel_file_path}")